import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from pathlib import Path
from datetime import datetime

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
REPORTS_DIR = Path("reports")

DATA_START_ROW = 4
DEFAULT_DATA_END_ROW = 37
_IDS_SHEET = "_ids"
_END_ROW_CELL = "B1"
MAX_EXPAND = 100


class DuplicateReceiptError(Exception):
    pass


class ReportFullError(Exception):
    def __init__(self, current_size: int):
        self.current_size = current_size
        super().__init__(f"Report is full ({current_size} rows)")


def get_user_report_path(user_id: int) -> Path:
    """Путь к файлу-реестру конкретного пользователя за текущий месяц."""
    return REPORTS_DIR / f"{user_id}_{datetime.now():%Y-%m}.xlsx"


def _get_end_row(wb) -> int:
    if _IDS_SHEET not in wb.sheetnames:
        return DEFAULT_DATA_END_ROW
    val = wb[_IDS_SHEET][_END_ROW_CELL].value
    return int(val) if val else DEFAULT_DATA_END_ROW


def _set_end_row(wb, end_row: int) -> None:
    ws = _get_or_create_ids_sheet(wb)
    ws[_END_ROW_CELL].value = end_row


def _get_or_create_ids_sheet(wb):
    if _IDS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_IDS_SHEET)
        ws.sheet_state = "hidden"
        return ws
    return wb[_IDS_SHEET]


def _next_empty_row(ws, end_row: int) -> int | None:
    for row_idx in range(DATA_START_ROW, end_row + 1):
        if ws.cell(row=row_idx, column=2).value is None:
            return row_idx
    return None


def _get_seen_ids(wb) -> set[str]:
    if _IDS_SHEET not in wb.sheetnames:
        return set()
    # Receipt IDs живут в колонке A, начиная с A2 (B1 занята под end_row).
    return {
        str(row[0].value)
        for row in wb[_IDS_SHEET].iter_rows(min_row=2, max_col=1)
        if row[0].value
    }


def _record_id(wb, receipt_id: str) -> None:
    ws = _get_or_create_ids_sheet(wb)
    next_row = ws.max_row + 1 if ws.cell(row=ws.max_row, column=1).value else ws.max_row
    if next_row < 2:
        next_row = 2
    ws.cell(row=next_row, column=1).value = receipt_id


def get_or_create_report(report_path: Path) -> Path:
    if not report_path.exists():
        REPORTS_DIR.mkdir(exist_ok=True)
        shutil.copy(TEMPLATE_PATH, report_path)
    return report_path


def add_receipt(
    user_id: int,
    org_name: str,
    amount: float,
    vat: float,
    payment_date: datetime,
    receipt_id: str,
) -> Path:
    path = get_user_report_path(user_id)
    get_or_create_report(path)

    wb = load_workbook(path)
    ws = wb.active

    if receipt_id in _get_seen_ids(wb):
        raise DuplicateReceiptError(receipt_id)

    end_row = _get_end_row(wb)
    row = _next_empty_row(ws, end_row)
    if row is None:
        raise ReportFullError(end_row - DATA_START_ROW + 1)

    ws.cell(row=row, column=2).value = org_name
    ws.cell(row=row, column=3).value = amount
    ws.cell(row=row, column=4).value = vat
    ws.cell(row=row, column=5).value = payment_date

    _record_id(wb, receipt_id)
    wb.save(path)
    return path


def expand_report(user_id: int, n: int) -> int:
    """Добавляет n пустых строк к таблице, сдвигая ИТОГО и подпись вниз.
    Возвращает новый DATA_END_ROW."""
    if not 1 <= n <= MAX_EXPAND:
        raise ValueError(f"n должен быть от 1 до {MAX_EXPAND}")

    path = get_user_report_path(user_id)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path)
    ws = wb.active

    old_end = _get_end_row(wb)
    insert_at = old_end + 1  # вставляем перед бывшей строкой 38 (пустой разделитель)
    template_row = old_end

    # openpyxl.insert_rows не сдвигает merged ranges — делаем сами.
    ranges_to_shift = [
        str(mr) for mr in ws.merged_cells.ranges if mr.min_row >= insert_at
    ]
    for r in ranges_to_shift:
        ws.unmerge_cells(r)

    ws.insert_rows(insert_at, amount=n)

    for r in ranges_to_shift:
        min_col, min_row, max_col, max_row = range_boundaries(r)
        shifted = (
            f"{get_column_letter(min_col)}{min_row + n}:"
            f"{get_column_letter(max_col)}{max_row + n}"
        )
        ws.merge_cells(shifted)

    # Копируем стили и автонумерацию из последней строки данных.
    last_number = ws.cell(row=template_row, column=1).value or (template_row - DATA_START_ROW + 1)
    try:
        last_number = int(last_number)
    except (TypeError, ValueError):
        last_number = template_row - DATA_START_ROW + 1

    for offset in range(n):
        new_row = old_end + 1 + offset
        for col in range(1, 6):
            src = ws.cell(row=template_row, column=col)
            dst = ws.cell(row=new_row, column=col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
        ws.cell(row=new_row, column=1).value = last_number + 1 + offset

    new_end = old_end + n

    # Переписываем SUM-формулу под новый диапазон (формула уехала вниз вместе с ИТОГО).
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=SUM(D"):
                cell.value = f"=SUM(D{DATA_START_ROW}:D{new_end})"

    _set_end_row(wb, new_end)
    wb.save(path)
    return new_end


def clear_user_reports(user_id: int) -> int:
    """Удаляет все файлы-реестры пользователя. Возвращает кол-во удалённых файлов."""
    if not REPORTS_DIR.exists():
        return 0
    deleted = 0
    for f in REPORTS_DIR.glob(f"{user_id}_*.xlsx"):
        f.unlink()
        deleted += 1
    return deleted
